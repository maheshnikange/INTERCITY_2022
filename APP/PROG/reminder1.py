import PyPDF2
import pandas as pd
import openpyxl
import datetime
from fpdf import FPDF
from datetime import datetime
from datetime import datetime
from barcode import EAN13,Code39
from barcode.writer import ImageWriter
from fpdf import FPDF
from PIL import Image
from PIL import Image
from APP.STD_FILES.postcodefile import STD_GROUP_CODE
from APP.STD_FILES.logfile import Log_File
from .bkrcc import *
from APP.models import *

def rotate_func(kk):
    pdf_in = open(r"%s\rem11.pdf"%kk, 'rb')
    pdf_reader = PyPDF2.PdfFileReader(pdf_in)
    pdf_writer = PyPDF2.PdfFileWriter()

    for pagenum in range(pdf_reader.numPages):
        page = pdf_reader.getPage(pagenum)
        if pagenum % 2:
            page.rotateClockwise(180)
        pdf_writer.addPage(page)

    pdf_out = open(r"%s\rem1.pdf"%kk, 'wb')
    pdf_writer.write(pdf_out)
    pdf_out.close()
    pdf_in.close() 
    
    import os
    os.remove(r"%s\rem11.pdf"%kk)


def REMINDER1(request,files1):
    move_files = []
    move_files.append(files1)
    p = 1
    cppp = 1
    def Folder_creation(cppp):
        # ------------------Capture date----------------
        cppp=str(cppp)
        today=date.today()
        d1 = today.strftime("%d-%b-%Y")
        # print(d1,'ddddddddddddddddddddddddddddddddd')
        d2=d1[0:2]
        d3=d1[3:6]
        d1=d2+d3+cppp
        # ------------------Capture time----------------
        now = datetime.now()        
        start_time = now.strftime("%H:%M:%S")
        s1=start_time[0:2]
        s2=start_time[3:5]
        s3=start_time[6:8]
        start_time=s1+s2+s3

        # ---------------create first date folder------------
        # import os
        # directory = d
        # print('Helloooooooooooooooooooooooooooooooo')
        # directory=d1
        # parent_dir = "E:\IWOC\OUTPUT\Bank Simpanan Nasional\CYCLE\\"
        outp = Scheduler.objects.filter(PROGRAM_NAME='REMINDER1').values()
        res = []
        otp = []
        for idx, sub in enumerate(outp, start = 0):
            if idx == 0:
                res.append(list(sub.values()))
        new = [i[-1] for i in res]
        
        for z in new:
            directory=d1
            # parent_dir = "E:\IWOC\OUTPUT\Bank Simpanan Nasional\CYCLE\\"
            parent_dir = z
            directory = d1+"_"+start_time 
            path1 = os.path.join(parent_dir, directory) # date folder path
            os.mkdir(path1)
            # print("Directory '% s' created" % directory)

            directory = 'BsnReminder1'
            parent_dir = path1
            path2 = os.path.join(path1, directory) # bkrcc folder path
            os.mkdir(path2)
            # print("Directory '% s' created" % directory)

            directory = 'PRINTING'
            parent_dir = path2
            path3 = os.path.join(path2, directory) # DIGITAL folder path
            os.mkdir(path3)
            # print("Directory '% s' created" % directory)

            # directory = 'PRINTING'
            # parent_dir = path2
            # path4 = os.path.join(path2, directory) # PRINTING folder path
            # os.mkdir(path4)
            # print("Directory '% s' created" % directory)

            return path3

    path11=r'E:\IWOC\SOURCE\\'+ files1
    data=pd.read_excel(path11)
    # print(data.columns)
    df=pd.DataFrame()
    df['br_code']=data['ILOM_BRCH_CODE']
    df['CUSTOMER_NAME']=data['CUSTOMER_NAME']
    df['ADDR1']=data['ADDR1']
    df['ADDR2']=data['ADDR2']
    df['ADDR3']=data['ADDR3']
    df['ADDR4']=data['ADDR4']
    df['POSTCODE']=data['POSTCODE']
    df['CITY_CD']=data['CITY_CD']
    df['STATE']=data['STATE']
    df['ILOM_OLD_SEQUENCE']=data['ILOM_OLD_SEQUENCE']
    df['ICBS_PROD_SUBDESC']=data['ICBS_PROD_SUBDESC']
    df['TOTAL_DUE']=data['TOTAL_DUE']
    kk1=data['SYSDATE_GENERATE'].to_list()
    List_of_Postcodes=data['POSTCODE'].to_list() 

    amt = data['ILOM_BRCH_CODE'].to_list()
    fi1 = []
    for i in amt:
        i = str(i)
        if len(i) == 4:
            res = '0' + str(i)
            fi1.append(res)
        else:
            fi1.append(i)
    
    date0=[]
    date2 = []
    for i in kk1:
        date0.append(i.strftime("%d/%m/%Y"))
        date2.append(i.strftime("%d%m%y"))
    df['SYSDATE_GENERATE']=date0
    # print(df['SYSDATE_GENERATE'],'dateeeee')

    received_grp_code= STD_GROUP_CODE(List_of_Postcodes)
    df['Group_code']=received_grp_code


                                                                                 
    time1 = datetime.now() 
    start_time=time1.strftime("%H:%M:%S")
    date1 = time1.strftime("%d/%b/%Y")
    total_clients=len(List_of_Postcodes)
    
    
    #--------------------------function for barcode------------------------------

    def barcode(acc_no,count):
        now = datetime.now() 
        date = now.strftime("%m%d%y")
        barcode_no = '0901' +  str(date) +  str(acc_no) 
        # print("date :",date) 

        number = barcode_no
        my_code = Code39(number, writer=ImageWriter())
        new_code  = 'b' + str(count)
        my_code.save(r"E:\IWOC\EXTRA\barcode_for_rem1\%s" %(new_code))

        from PIL import Image

        img = Image.open(r"E:\IWOC\EXTRA\barcode_for_rem1\%s.png" %(new_code))
        imgCropped = img.crop(box=(4,10,1100,200))
        imgCropped.save(r"E:\IWOC\EXTRA\barcode_for_rem1\%s.png" %(new_code))

    count = 1
    c = 1
    df['Group_code1'] = df['Group_code'].str[1:]
    final_df = df.sort_values(by = 'Group_code1',ignore_index = True)
        
    pdf = FPDF()
    for i in range(0,total_clients):
        # print(final_df['Group_code1'][i],'-===============1111111111')
        barcode(final_df['ILOM_OLD_SEQUENCE'][i],count)
        pdf.add_page()
        pdf.set_auto_page_break(auto=False)
        pdf.set_line_width(6)
        pdf.set_draw_color(r=0, g=0, b=0)
        pdf.line(x1=5, y1=4, x2=205, y2=4)         #1st horizontal line
        pdf.line(x1=5, y1=292, x2=205, y2=292)     #2nd horizontal line 
        pdf.line(5,6,5,290)                         # 1st vertical line
        pdf.line(205,6,205,290)                     # 2nd vertical line    

        pdf.set_font('Times','B',8.5)
        pdf.cell(45)
        pdf.cell(10,10,txt = 'PERATURAN-PERATURAN PEMBAYARAN PINJAMAN/PEMBIAYAAN')
        pdf.ln(1)
        pdf.cell(66)
        pdf.cell(10,14,txt = 'BANK SIMPANAN NASIONAL (BSN)')

        pdf.ln(2)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,30,txt ='Bayaran PINJAMAN / PEMBIAYAAN boleh dibuat secara tunai di mana-mana ATM/CDM/Cawangan BSN, pindahan GIRO / GIRO-I, Perbankan internet') 
        pdf.ln(1)
        pdf.cell(10,36, txt = '(www.mybsn.com.my) dan Inter Bank Giro (IBG) dari bank-bank lain yang mengambil bahagian.')


        pdf.ln(4)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,44,txt ='Bayaran PINJAMAN/PEMBIAYAAN juga boleh dibuat menerusi Arahan Tetap akaun simpanan GIRO/GIRO-I. Bagi membolehkan kami mendebit Akaun') 
        pdf.ln(1)
        pdf.cell(10,50, txt = 'Simpanan BSN tuan/puan, sila pastikan amaun tuan/puan mempunyai baki minima yang ditetapkan oleh BSN dan amaun yang perlu dijelaskan.')


        pdf.ln(4)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,58,txt ='Sekiranya pembayaran melalui cek, draf bank atau arahan juruwang hendaklah dibuat atas nama') 
        pdf.cell(106)
        pdf.set_font('Times','B',8.5)
        pdf.cell(10,58,txt ='BANK SIMPANAN NASIONAL') 
        pdf.cell(33)
        pdf.set_font('Times','',8.5)
        pdf.cell(10,58,txt ='dan dipalang') 
        pdf.cell(5.8)
        pdf.set_font('Times','B',8.5)
        pdf.cell(10,58,txt =' "AKAUN') 
        pdf.ln(1)
        pdf.set_font('Times','B',8.5)
        pdf.cell(10,64, txt = 'PENERIMA SAHAJA"')
        pdf.set_font('Times','',8.4)
        pdf.cell(20)
        pdf.cell(10,64, txt = '. Catatkan nama dan nombor akaun Pinjaman / Pembiayaan tuan/puan di belakang cek berkenaan. Bayaran hendaklah sampai kepada')
        pdf.ln(1)
        pdf.cell(10,70, txt = 'kami sebelum tarikh akhir yang ditetapkan.')


        pdf.ln(4)
        pdf.set_font('Times','',9)
        pdf.cell(10,76,txt ='Denda lewat bayar / caj perkhidmatan akan dikenakan jika bayaran yang dibuat oleh tuan / puan diterima selepas tarikh akhir  yang ditetapkan. Jika') 
        pdf.ln(1)
        pdf.cell(10,82, txt = 'tuan/puan telah menjelaskan amaun berkenaan, sila abaikan pemberitahuan ini.')

        pdf.ln(4)
        pdf.set_font('Times','',9)
        pdf.cell(10,88,txt ='Jika tuan / puan memerlukan penjelasan selanjutnya berhubung butir / butir yang dinyatakan di hadapan, sila hubungi  kami di  talian 03-2028 3222') 
        pdf.ln(1)
        pdf.cell(10,94, txt = 'atau cawangan di mana tuan / puan membuat permohonan pinjaman / pembiayaan tersebut di talian-talian berikut:-')

        pdf.ln(1)
        pdf.cell(10,104,txt = 'Johor')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Kedah')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Kelantan')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'N. Sembilan')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Melaka')
        pdf.cell(15)
        pdf.cell(10,104,txt = 'Pahang')

        pdf.ln(1)
        pdf.cell(10,108,txt = '07-208 3555')
        pdf.cell(15)
        pdf.cell(10,108,txt = '04-7740 444')
        pdf.cell(15)
        pdf.cell(10,108,txt = '09-745 7070')
        pdf.cell(15)
        pdf.cell(10,108,txt = '06-768 6500')
        pdf.cell(15)
        pdf.cell(10,108,txt = '06-289 5800')
        pdf.cell(15)
        pdf.cell(10,108,txt = '09-565 0565')


        pdf.ln(4)
        pdf.cell(10,110,txt = 'Perak')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'P.Pinang')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Sabah')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Sarawak')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Terengganu')
        pdf.cell(15)
        pdf.cell(10,110,txt = 'Selangor')

        pdf.ln(1)
        pdf.cell(10,114,txt = '05-245 2222')
        pdf.cell(15)
        pdf.cell(10,114,txt = '04-222 6400')
        pdf.cell(15)
        pdf.cell(10,114,txt = '088-355 600')
        pdf.cell(15)
        pdf.cell(10,114,txt = '082-227 888')
        pdf.cell(15)
        pdf.cell(10,114,txt = '09-6200 400')
        pdf.cell(15)
        pdf.cell(10,114,txt = '03-5543 3000')

        pdf.ln(1)
        pdf.set_line_width(1)
        pdf.line(10,105,200,105)

        pdf.ln(1)
        pdf.set_font('Arial','',8)
        pdf.cell(90)
        pdf.cell(10,142,txt = 'SULIT')

        pdf.ln(1)
        pdf.cell(10,146, txt = 'JABATAN PEMULIHAN KREDIT')
        pdf.cell(140)
        pdf.cell(10,146,txt = 'TANPA PRASANGKA')

        pdf.ln(1)
        pdf.cell(10,152,txt = 'BANK SIMPANAN NASIONAL')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,158,txt ="(")
        pdf.cell(-8)
        pdf.cell(10,158,txt =  str(fi1[i]))
        pdf.cell(-1)
        pdf.cell(10,158,txt =")")


        pdf.ln(1)
        pdf.cell(160)
        pdf.cell(10,164,txt = str(final_df['SYSDATE_GENERATE'][i]))

        h = 170
        if pd.notna(final_df['CUSTOMER_NAME'][i]):
            pdf.ln(1)
            pdf.cell(60)
            # print(i,final_df['CUSTOMER_NAME'][i],'---2')
            pdf.cell(10,h,txt = str(final_df['CUSTOMER_NAME'][i]))

        if pd.notna(final_df['ADDR1'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(final_df['ADDR1'][i]))

        if pd.notna(final_df['ADDR2'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(final_df['ADDR2'][i]))

        if pd.notna(final_df['ADDR3'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(final_df['ADDR3'][i]))

        if pd.notna(final_df['ADDR4'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(final_df['ADDR4'][i]))

        if pd.notna(final_df['POSTCODE'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(final_df['POSTCODE'][i] ) )
            pdf.cell(-1)
            pdf.cell(10,h,txt=  str(final_df['CITY_CD'][i]))

        if pd.notna(final_df['STATE'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(60)
            pdf.cell(10,h,txt = str(final_df['STATE'][i]))

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,220, txt = 'Nombor Akaun :')

        pdf.cell(12)
        pdf.cell(10,220, txt = str(final_df['ILOM_OLD_SEQUENCE'][i]))


        pdf.cell(80)
        pdf.cell(10,220, txt = str(final_df['ICBS_PROD_SUBDESC'][i]))


        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,250,txt = 'PERINGATAN 1 ')

        pdf.ln(1)
        pdf.cell(10)
        pdf.cell(10,256,txt = 'NOTIS TUNTUTAN TUNGGAKAN PEMBIAYAAN / PINJAMAN ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,270,txt = 'Butir - butir pembiayaan / pinjaman Tuan / Puan sehingga tarikh notis tuntutan ini  menunjukkan tunggakan sebanyak')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        a = str(final_df['TOTAL_DUE'][i])
        a = str(a)
        n = len(a)
        if n > 7:
            a = a[0:1] + ',' +  a[1:n] 
            pdf.cell(10,276,txt = 'RM ' + str(a) + '.')
        elif n >= 7:
            a = a[0:1] + ',' +  a[1:n] 
            pdf.cell(10,276,txt = 'RM ' + str(a) + '.')
        else:
            pdf.cell(10,276,txt = 'RM ' + str(final_df['TOTAL_DUE'][i]) + '.')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9.5)
        pdf.cell(10,290,txt = 'Sila jelaskan  amaun tunggakan  dalam tempoh 14 hari dari tarikh notis ini. Hubungi kami melalui talian  seperti')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,296,txt = 'lampiran di atas (Seksyen Kutipan) jika Tuan / Puan memerlukan penjelasan lanjut. ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9.5)
        pdf.cell(10,310,txt = 'Hubungi  Agensi  Kaunseling  dan  Pengurusan  Kredit (AKPK)  yang  menyediakan  perkhidmatan  pengurusan ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9.4)
        pdf.cell(10,316,txt = 'kewangan, kaunseling kredit, pendidikan kewangan dan penstrukturan semula pinjaman secara percuma kepada ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9.2)
        pdf.cell(10,322,txt = 'individu. Untuk sebarang pertanyaan, sila hubungi talian 1-800-88-2575. ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,348,txt = 'Ingin bayar bil dan semak baki  akaun simpanan anda secara online? Daftarlah Perbankan Internet MyBSN sekarang! ')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9.3)
        pdf.cell(10,354,txt = 'Kemudahan perbankan hanya untuk anda dari BSN. Untuk maklumat lanjut, layari www.mybsn.com.my atau 1300')


        pdf.ln(1)
        pdf.cell(10)
        pdf.set_font('Arial','',9)
        pdf.cell(10,360,txt = '88 1900.')

        pdf.ln(1)
        pdf.cell(140)
        pdf.set_font('Arial','',5)
        pdf.cell(10,410,txt = 'R1 21112022-24112022_2')

        pdf.cell(12)
        pdf.set_font('Arial','',5)
        pdf.cell(10,410,txt = '00000' + str(count))
        

        pdf.add_page()
        pdf.image(r'E:\IWOC\APP SCRIPT\REMINDER1\CYCLE\REMINDER1\DATA\BSN_Reminder-01.jpg',-10,-10,230)

        pdf.set_font('Arial','',7)
        pdf.cell(10,280,txt = '00' + str(c))
        c += 1
        pdf.ln(1)
        pdf.image(r'E:\IWOC\APP SCRIPT\REMINDER1\CYCLE\REMINDER1\DATA\BSN_Reminder-02.jpg',20,146,80)
        pdf.image(r'E:\IWOC\APP SCRIPT\REMINDER1\CYCLE\REMINDER1\DATA\BSN_Reminder-02 - Copy.jpg', 145,146,40)

        pdf.ln(1)
        new_code  = 'b' + str(count)
        pdf.image(r'E:\IWOC\EXTRA\barcode_for_rem1\%s.png'%(new_code),35,214,40)  # 70
        

        #--------------------------------image crop-------------------------------------
       
        
        pdf.ln(1)
        pdf.set_font('times','',7)
        pdf.cell(25)
        count  = str(count)
        if len(count) == 1:
            count_value = '00000' + str(count)
        if len(count) == 2:
            count_value = '0000' + str(count)
        if len(count) == 3:
            count_value = '000' + str(count)
        if len(count) == 4:
            count_value = '00' + str(count)
        if len(count) == 5:
            count_value = '0' + str(count)
        if len(count) == 6:
            count_value =  str(count)
        pdf.cell(10,419, txt  = str(count_value))
        pdf.cell(-2)
        pdf.cell(10,419, txt = "-" + str(date2[i]))
        pdf.cell(5)
        pdf.cell(10,419,txt = '(' + str(final_df['Group_code'][i])+  ')' + '-P0 1-01')
        count = int(count)
        count += 1

        h = 430
        if pd.notna(final_df['CUSTOMER_NAME'][i]):
            pdf.ln(1)
            pdf.set_font('Arial','',10)
            pdf.cell(25)
            pdf.cell(10,h,txt = str(final_df['CUSTOMER_NAME'][i]))

        if pd.notna(final_df['ADDR1'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(25)
            pdf.cell(10,h,txt = str(final_df['ADDR1'][i]))

        if pd.notna(final_df['ADDR2'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(25)
            pdf.cell(10,h,txt = str(final_df['ADDR2'][i]))

        if pd.notna(final_df['ADDR3'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(25)
            pdf.cell(10,h,txt = str(final_df['ADDR3'][i]))

        if pd.notna(final_df['ADDR4'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(25)
            pdf.cell(10,h,txt= str(final_df['ADDR4'][i]))

        if pd.notna(final_df['POSTCODE'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(25)
            pdf.cell(10,h,txt = str(final_df['POSTCODE'][i]))
            pdf.cell(1)
            pdf.cell(10,h,txt = str(final_df['CITY_CD'][i]))

        if pd.notna(final_df['STATE'][i]):
            h = h + 6
            pdf.ln(1)
            pdf.cell(25)
            pdf.cell(10,h,txt = str(final_df['STATE'][i]))

    kk=Folder_creation(cppp)
    bk = kk
        
    pdf.output(r"%s\rem11.pdf"%kk)
    print('completed')
    if i == total_clients-1:
        print(i,'iii')
        rotate_func(bk)
    # lesson =(r"%s\rem1.pdf"%kk)
    # file_name = os.path.split(lesson)
    # ffname = file_name[-1]
    # ffname = ffname.split('.')
    # ffname = ffname[0]

    import PyPDF2
    file = open(r"%s\rem1.pdf"%kk, 'rb')
    readpdf = PyPDF2.PdfFileReader(file)
    totalpages = readpdf.numPages
    cppp += 1
    
    
    end_time=time1.strftime("%H:%M:%S")
    
    path20 = r'E:\IWOC\SOURCE' 
    target = r'E:\IWOC\TEMP' 

    for i in move_files:

        src_path = os.path.join(path20, i)
        dst_path = os.path.join(target, i)
        os.rename(src_path, dst_path)
    print('Files Moved................')

    today = date.today()    
    user = DetailActivity.objects.last()
    first_record = final_df['ILOM_OLD_SEQUENCE'][0]
    last_record = final_df['ILOM_OLD_SEQUENCE'].iat[-1]
    ffname = 'reminder1'
    # data = History.objects.create(Date = str(today),Start = str(start_time),Finish=str(end_time),User = str(user) ,Product = 'KKK', File= str(files1),Status = 'Enable',Log = str(count),ActiveCount = str(total_clients),Page = str(total_clients),Impression  = 'DONE',filenamepdf = 'PGI.pdf',totalaccountspdf = str(total_clients),totalpagespdf =str(totalpages), firstrecordpdf = str(data['acc_no'][0]),lastrecordpdf = str(data['acc_no'].iat[-1]), programnamepdf = 'PGI' )                                                                                                                                                                                                           
    Log_File(today,start_time,end_time,user,ffname,files1,total_clients,totalpages,first_record,last_record)
    # Log_File(path11)




